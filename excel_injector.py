import openpyxl
from openpyxl.styles import Font
from copy import copy
import io
import os

from config_referensi import BLACKLIST_CLUSTER, BLACKLIST_SUBFEEDER

# =============================================================================
# MODUL 1: UTILITAS & PERLINDUNGAN
# =============================================================================
def load_external_placeholders(filepath="placeholders.txt"):
    mapping = {}
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    mapping[key.strip()] = val.strip()
    return mapping

def is_protected_sheet(ws):
    """Mengamankan sheet dengan tab warna merah (#FF0000) agar tidak tersentuh pemusnahan/modifikasi"""
    tab_color = ws.sheet_properties.tabColor
    if tab_color and tab_color.rgb:
        if 'FF0000' in tab_color.rgb.upper():
            return True
    return False

def apply_black_font(cell):
    """Mengunci warna font menjadi hitam pekat (#000000) tanpa merusak ukuran template"""
    if cell.has_style and cell.font:
        new_font = copy(cell.font)
        new_font.color = "000000"
        cell.font = new_font

# =============================================================================
# MODUL 2: MESIN SPILLOVER (KIRI-BAWAH -> KANAN)
# =============================================================================
def process_ba_splitter_spillover(wb, template_prefixes, title_base, fat_items, placeholder="[INPUT_FAT_NAME]"):
    """
    Khusus BA Splitter: Jika slot baris di tabel tidak dipakai, SELURUH teks di baris itu dihapus.
    Sehingga tulisan "Splitter 10", "MICRO", dsb hilang menyisakan border kosong.
    """
    available_sheets = []
    for pref in template_prefixes:
        available_sheets.extend([s for s in wb.sheetnames if s.startswith(pref)])
    available_sheets = sorted(available_sheets)
    
    item_idx = 0
    sheet_count = 1
    
    for s_name in available_sheets:
        if item_idx >= len(fat_items):
            break
            
        ws = wb[s_name]
        ws.title = f"{title_base}" if sheet_count == 1 else f"{title_base}_{sheet_count}"
        
        found_rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and placeholder in cell.value:
                    found_rows.append(row)
                    break
        
        for row in found_rows:
            if item_idx < len(fat_items):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and placeholder in cell.value:
                        cell.value = cell.value.replace(placeholder, fat_items[item_idx])
                        apply_black_font(cell)
                item_idx += 1
            else:
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = ""
                        
        sheet_count += 1

def process_spillover_dynamic(wb, template_prefixes, title_base, items, placeholder="[INPUT_FAT_NAME]"):
    """Spillover standar untuk OTDR dan Distribusi. Hanya menghapus tag placeholdernya saja jika tidak terpakai."""
    available_sheets = []
    for pref in template_prefixes:
        available_sheets.extend([s for s in wb.sheetnames if s.startswith(pref)])
    available_sheets = sorted(available_sheets)
    
    item_idx = 0
    sheet_count = 1
    
    for s_name in available_sheets:
        if item_idx >= len(items):
            break
            
        ws = wb[s_name]
        new_title = f"{title_base}" if sheet_count == 1 else f"{title_base}_{sheet_count}"
        ws.title = new_title[:31]
        
        cells = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and placeholder in cell.value:
                    cells.append(cell)
                    
        cells.sort(key=lambda c: (c.column, c.row))
        
        for cell in cells:
            if item_idx < len(items):
                cell.value = cell.value.replace(placeholder, items[item_idx])
                apply_black_font(cell)
                item_idx += 1
            else:
                cell.value = cell.value.replace(placeholder, "")
                
        sheet_count += 1

# =============================================================================
# MODUL 3: INJEKTOR UTAMA (FASE 1 & FASE 2)
# =============================================================================
def inject_excel_final(template_stream, metadata, parsed_fat, parsed_poles, mode="cluster"):
    wb = openpyxl.load_workbook(template_stream)
    p_map = load_external_placeholders()
    
    fat_by_line = {}
    for fat in parsed_fat:
        line_letter = "".join(filter(str.isalpha, fat)) or "A"
        if line_letter not in fat_by_line:
            fat_by_line[line_letter] = []
        fat_by_line[line_letter].append(fat)
    line_keys = sorted(list(fat_by_line.keys()))

    # --- 1. FAT INDIVIDU ---
    template_fats = sorted([s for s in wb.sheetnames if s.startswith("FAT_")])
    if mode == "subfeeder":
        if template_fats and parsed_fat:
            ws = wb[template_fats[0]]
            ws.title = "FAT"
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "[INPUT_FAT_NAME]" in cell.value:
                        cell.value = cell.value.replace("[INPUT_FAT_NAME]", f"FAT {parsed_fat[0]}")
                        apply_black_font(cell)
    else:
        for i, fat_code in enumerate(parsed_fat):
            if i < len(template_fats):
                ws = wb[template_fats[i]]
                ws.title = f"FAT {fat_code}"
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and "[INPUT_FAT_NAME]" in cell.value:
                            cell.value = cell.value.replace("[INPUT_FAT_NAME]", f"FAT {fat_code}")
                            apply_black_font(cell)

    # --- 2. BA SPLITTER FAT (Penghapusan Baris Bersih) ---
    for line in line_keys:
        fat_names = [f"FAT {f}" for f in fat_by_line[line]]
        process_ba_splitter_spillover(wb, ["BA Spitter FAT_", "BA Splitter FAT_"], f"BA Splitter FAT {line}", fat_names)

    # --- 3. OTDR SUMMARY FDT-FAT ---
    for line in line_keys:
        fat_names = [f"FAT {f}" for f in fat_by_line[line]]
        process_spillover_dynamic(wb, ["OTDR Sumary (FDT-FAT)_", "master otdr summary"], f"OTDR Sumary (FDT-FAT) line {line}", fat_names)

    # --- 4. E2E OPM DISTRIBUTION ---
    for line in line_keys:
        distribution_names = [f"Line {line} No {str(idx+1).zfill(2)}" for idx in range(len(fat_by_line[line]))]
        title_dist = "E2E OPM Distribution" if line == "A" else f"E2E OPM Distribution {line}"
        process_spillover_dynamic(wb, ["OPM_DISTRIBUTION_"], title_dist, distribution_names)

    # --- 5. E2E OPM FEEDER / SUBFEEDER ---
    template_opm = sorted([s for s in wb.sheetnames if s.startswith("OPM_") and "DIST" not in s.upper()])
    if template_opm:
        wb[template_opm[0]].title = "E2E OPM Feeder" if mode == "cluster" else "E2E OPM Subfeeder"

    # --- 6. OTDR SUMMARY WAVE (1550 & 1310) ---
    template_otdr_wave = sorted([s for s in wb.sheetnames if s.startswith("OTDR Summary (WAVE)_")])
    for i, s_name in enumerate(template_otdr_wave):
        if i == 0: wb[s_name].title = "OTDR Summary 1550"
        elif i == 1: wb[s_name].title = "OTDR Summary 1310"

    # --- 7. POLE ERECTION (SPILLOVER MURNI TANPA INSERT ROW) ---
    available_poles = sorted([s for s in wb.sheetnames if s.startswith("POLE_")])
    for pole in parsed_poles:
        qty = pole["qty"]
        if qty <= 0: continue
        
        written = 0
        sheet_count = 1
        prefix = "P" if "NEW" in pole["type"] else "E"
        
        while written < qty and available_poles:
            s_name = available_poles.pop(0)
            ws = wb[s_name]
            title_base = pole["title"][:28]
            ws.title = f"{title_base}" if sheet_count == 1 else f"{title_base}_{sheet_count}"
            
            cells = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "[INPUT_POLE_DESC]" in cell.value:
                        cells.append(cell)
            cells.sort(key=lambda c: (c.column, c.row))
            
            for cell in cells:
                r = cell.row
                if written < qty:
                    serial = f"{prefix}{str(written+1).zfill(3)}"
                    cell.value = cell.value.replace("[INPUT_POLE_DESC]", f"{pole['type']} {serial}")
                    apply_black_font(cell)
                    
                    for c_iter in ws[r]:
                        if c_iter.value and isinstance(c_iter.value, str):
                            if "[INPUT_POLE_SIZE]" in c_iter.value:
                                c_iter.value = c_iter.value.replace("[INPUT_POLE_SIZE]", pole["size_clean"])
                                apply_black_font(c_iter)
                            elif "pass" in c_iter.value.lower() and "fail" in c_iter.value.lower():
                                c_iter.value = "□pass  □fail □Need Repair"
                                apply_black_font(c_iter)
                    written += 1
                else:
                    cell.value = cell.value.replace("[INPUT_POLE_DESC]", "")
                    for c_iter in ws[r]:
                        if c_iter.value and isinstance(c_iter.value, str):
                            if "[INPUT_POLE_SIZE]" in c_iter.value:
                                c_iter.value = c_iter.value.replace("[INPUT_POLE_SIZE]", "")
                            elif "pass" in c_iter.value.lower() and "fail" in c_iter.value.lower():
                                c_iter.value = ""
            sheet_count += 1

    # --- 8. METADATA REPLACEMENT ---
    # Di Fase 1: Hanya data identitas yang diganti, [DISTANCE] dsb tetap aman.
    # Di Fase 2: Seluruh angka pengukuran akan ditimpa ke dalam placeholder sisa.
    for sheet in wb.worksheets:
        if is_protected_sheet(sheet) or "cover" in sheet.title.lower():
            continue
            
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original_val = cell.value
                    changed = False
                    for key, val in metadata.items():
                        token = p_map.get(key, f"[{key}]") 
                        if token in original_val:
                            original_val = original_val.replace(token, str(val))
                            changed = True
                    if changed:
                        cell.value = original_val
                        apply_black_font(cell)

    # =============================================================================
    # MODUL 4: PEMBERSIHAN FINAL (GARBAGE COLLECTOR)
    # =============================================================================
    sheets_to_remove = []
    for s_name in wb.sheetnames:
        ws = wb[s_name]
        if is_protected_sheet(ws):
            continue
            
        name_lower = s_name.lower().strip()
        if s_name.startswith(("FAT_", "POLE_", "BA Sp", "OTDR Sum", "OPM_DIST", "OPM_", "OTDR Summary (WAVE)_")):
            sheets_to_remove.append(s_name)
            continue
            
        if mode == "cluster":
            for blacklist in BLACKLIST_CLUSTER:
                if blacklist in name_lower:
                    sheets_to_remove.append(s_name)
                    break
        elif mode == "subfeeder":
            for blacklist in BLACKLIST_SUBFEEDER:
                if blacklist in name_lower:
                    sheets_to_remove.append(s_name)
                    break
                    
    for s_name in set(sheets_to_remove):
        if s_name in wb.sheetnames:
            wb.remove(wb[s_name])

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream
